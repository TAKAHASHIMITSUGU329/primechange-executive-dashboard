from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter
from datetime import datetime
import calendar
from pathlib import Path
from typing import Any
from urllib.request import urlopen

from openpyxl import load_workbook


MASTER_SHEET_ID = "1f3DjbrRJifbIK1n-Xpfup-O5cFExnqr4G3cs75Aa43U"
MASTER_GID = "117684960"
MASTER_EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{MASTER_SHEET_ID}/export?format=csv&gid={MASTER_GID}"
)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
CURRENT_MONTH = "2026-03"
PREVIOUS_MONTH = "2026-02"
CURRENT_DATE = datetime(2026, 3, 13)

POSITIVE_CATEGORIES = {
    "清潔感": ["清潔", "きれい", "綺麗", "clean", "tidy", "well maintained", "清掃"],
    "スタッフ対応": ["スタッフ", "親切", "helpful", "friendly", "service", "接客"],
    "立地": ["立地", "駅近", "location", "station", "便利", "アクセス"],
    "朝食": ["朝食", "breakfast", "ビュッフェ"],
    "快適性": ["快適", "静か", "comfortable", "居心地", "quiet"],
}

NEGATIVE_CATEGORIES = {
    "水回り臭気": ["臭い", "におい", "匂い", "下水", "sewage", "smell", "odor", "odour"],
    "仕上がり不足": ["汚", "ほこり", "埃", "髪", "dirty", "not so clean", "unclean", "dust"],
    "薬剤・アレルギー": ["アレルギー", "allergy", "chemicals", "洗剤", "pests", "害虫"],
    "設備老朽化": ["古い", "old", "老朽", "古さ"],
    "騒音": ["騒音", "noise", "うるさい", "noisy"],
}

CLEANING_KEYWORDS = [
    "清潔",
    "きれい",
    "綺麗",
    "清掃",
    "汚",
    "ほこり",
    "埃",
    "髪",
    "臭い",
    "におい",
    "匂い",
    "下水",
    "sewage",
    "smell",
    "odour",
    "odor",
    "allergy",
    "アレルギー",
    "洗剤",
    "dirty",
    "clean",
    "tidy",
]


def fetch_bytes(url: str) -> bytes:
    with urlopen(url, timeout=20) as response:
        return response.read()


def fetch_csv_rows(url: str) -> list[list[str]]:
    text = fetch_bytes(url).decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def extract_sheet_id(url: str) -> str | None:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    return match.group(1) if match else None


def normalize_rating(source: str, raw_rating: Any) -> float | None:
    try:
        value = float(raw_rating)
    except (TypeError, ValueError):
        return None

    if source in {"Google", "じゃらん", "楽天トラベル"} or value <= 5:
        return value * 2
    return value


def categorize(text: str, dictionary: dict[str, list[str]]) -> str:
    lowered = text.lower()
    for category, keywords in dictionary.items():
        if any(keyword.lower() in lowered for keyword in keywords):
            return category
    return "未分類"


def is_cleaning_related(text: str) -> bool:
    lowered = text.lower()
    return any(keyword.lower() in lowered for keyword in CLEANING_KEYWORDS)


def find_review_header_index(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows):
        if "サイト名" in row and "評価" in row and "投稿日" in row:
            return index
    return -1


def parse_review_block(
    row: list[str], offset: int, hotel_name: str, seen: set[str], reviews: list[dict[str, Any]]
) -> None:
    source = (row[offset + 2] if len(row) > offset + 2 else "") or ""
    rating_raw = row[offset + 3] if len(row) > offset + 3 else ""
    review_date = row[offset + 4] if len(row) > offset + 4 else ""

    if not source or not rating_raw or not review_date:
        return

    full_comment = ((row[offset + 5] if len(row) > offset + 5 else "") or "").strip()
    positive_comment = ((row[offset + 6] if len(row) > offset + 6 else "") or "").strip()
    negative_comment = ((row[offset + 7] if len(row) > offset + 7 else "") or "").strip()
    translated_full = ((row[offset + 8] if len(row) > offset + 8 else "") or "").strip()
    translated_positive = ((row[offset + 9] if len(row) > offset + 9 else "") or "").strip()
    translated_negative = ((row[offset + 10] if len(row) > offset + 10 else "") or "").strip()
    author = ((row[offset + 11] if len(row) > offset + 11 else "") or "不明").strip()

    joined = " ".join(
        part
        for part in [
            full_comment,
            positive_comment,
            negative_comment,
            translated_full,
            translated_positive,
            translated_negative,
        ]
        if part
    )

    record = {
        "hotelName": hotel_name,
        "source": source.strip(),
        "ratingRaw": rating_raw,
        "normalizedRating": normalize_rating(source.strip(), rating_raw),
        "reviewDate": str(review_date).strip(),
        "fullComment": full_comment,
        "positiveComment": positive_comment,
        "negativeComment": negative_comment,
        "author": author,
        "positiveCategory": categorize(joined, POSITIVE_CATEGORIES),
        "negativeCategory": categorize(joined, NEGATIVE_CATEGORIES),
        "cleaningRelated": is_cleaning_related(joined),
    }

    key = "|".join(
        [
            record["source"],
            record["reviewDate"],
            record["author"],
            str(record["ratingRaw"]),
            record["fullComment"],
            record["positiveComment"],
            record["negativeComment"],
        ]
    )

    if key in seen:
        return

    seen.add(key)
    reviews.append(record)


def parse_reviews(rows: list[list[str]], hotel_name: str) -> list[dict[str, Any]]:
    header_index = find_review_header_index(rows)
    if header_index == -1:
        return []

    reviews: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row in rows[header_index + 1 :]:
        parse_review_block(row, 0, hotel_name, seen, reviews)
        parse_review_block(row, 12, hotel_name, seen, reviews)

    return reviews


def top_category(counter: Counter[str], allow_unclassified: bool = False) -> str:
    if not counter:
        return "特記事項なし"

    for category, _count in counter.most_common():
        if allow_unclassified or category != "未分類":
            return category
    return counter.most_common(1)[0][0]


def infer_region(hotel_name: str) -> str:
    if any(word in hotel_name for word in ["横浜", "川崎", "相模原"]):
        return "神奈川"
    if "成田" in hotel_name:
        return "千葉"
    if "博多" in hotel_name:
        return "福岡"
    return "東京"


def infer_risk_level(avg_review: float | None, cleaning_negative_rate: float) -> str:
    if avg_review is None:
        return "medium"
    if avg_review < 8 or cleaning_negative_rate >= 12:
        return "high"
    if avg_review < 8.6 or cleaning_negative_rate >= 7:
        return "medium"
    return "low"


def infer_upsell_level(review_count: int, cleaning_negative_rate: float, monthly_revenue: float | None) -> str:
    if review_count >= 20 and 4 <= cleaning_negative_rate < 12:
        return "high"
    if (monthly_revenue or 0) >= 1_500_000 and review_count >= 10:
        return "high"
    if review_count >= 10:
        return "medium"
    return "low"


def build_next_action(main_issue: str, risk_level: str, upsell_level: str) -> str:
    actions = {
        "水回り臭気": "臭気対策と水回り重点点検を提案",
        "仕上がり不足": "客室仕上がり監査と最終チェック強化を提案",
        "薬剤・アレルギー": "洗剤運用と寝具まわり点検の見直しを提案",
        "設備老朽化": "設備由来不満と清掃改善の切り分け報告を実施",
        "騒音": "清掃起因外のためホテル側課題として切り分け報告",
        "未分類": "レビュー内容の手動確認を優先",
    }
    if main_issue == "特記事項なし" and risk_level == "low":
        return "高評価運用を標準化して横展開"
    if upsell_level == "high" and risk_level == "low":
        return "高品質維持を前提に特別清掃メニューを提案"
    return actions.get(main_issue, "レビュー内容の手動確認を優先")


def latest_sheet_name(sheet_names: list[str], pattern: str) -> str | None:
    matched: list[tuple[int, str]] = []
    regex = re.compile(pattern)
    for name in sheet_names:
        hit = regex.search(name)
        if hit:
            matched.append((int(hit.group(1)), name))
    if not matched:
        return None
    return max(matched, key=lambda item: item[0])[1]


def to_float(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def row_values(ws, row_index: int, start_col: int, end_col: int) -> list[Any]:
    return [ws.cell(row=row_index, column=col).value for col in range(start_col, end_col + 1)]


def row_item(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if len(row) > index else None


def extract_financials(workbook, hotel_name: str) -> dict[str, Any]:
    sheet_names = workbook.sheetnames
    aggregate_sheet_name = latest_sheet_name(sheet_names, r"①R\d+_(\d+)集計")
    monthly_sheet_name = latest_sheet_name(sheet_names, r"④R\d+_(\d+)月報")
    annual_sheet_name = "年間集計" if "年間集計" in sheet_names else None
    special_sheet_name = "🧹特別清掃" if "🧹特別清掃" in sheet_names else None

    monthly_revenue = None
    previous_revenue = None
    target_revenue = None
    margin_pct = None
    previous_margin_pct = None
    target_margin_pct = None
    occupancy_pct = None
    previous_occupancy_pct = None
    variable_cost_rate = None
    net_profit = None
    previous_net_profit = None
    claim_rate = None
    rooms_cleaned = None
    days_in_month = calendar.monthrange(CURRENT_DATE.year, CURRENT_DATE.month)[1]
    special_cleaning_revenue = 0.0
    special_cleaning_cost = 0.0
    revenue_trend: list[dict[str, Any]] = []

    if aggregate_sheet_name:
        ws = workbook[aggregate_sheet_name]
        margin_pct = to_float(ws["C8"].value)
        target_margin_pct = to_float(ws["C7"].value)
        variable_cost_rate = to_float(ws["E8"].value)
        occupancy_pct = to_float(ws["G8"].value)
        monthly_revenue = to_float(ws["H8"].value)
        target_revenue = to_float(ws["H7"].value)
        net_profit = to_float(ws["L8"].value)
        rooms_cleaned = to_float(ws["G13"].value)

    if monthly_sheet_name:
        ws = workbook[monthly_sheet_name]
        claim_rate = to_float(ws["I10"].value)
        net_profit = net_profit if net_profit is not None else to_float(ws["H10"].value)

    if annual_sheet_name:
        ws = workbook[annual_sheet_name]
        month_cells = row_values(ws, 3, 3, 14)
        occupancy_cells = row_values(ws, 6, 3, 14)
        revenue_cells = row_values(ws, 7, 3, 14)
        net_profit_cells = row_values(ws, 13, 3, 14)
        margin_cells = row_values(ws, 15, 3, 14)
        for sheet_name, revenue in zip(month_cells, revenue_cells):
            if not sheet_name:
                continue
            month_match = re.search(r"①R\d+_(\d+)集計", str(sheet_name))
            if not month_match:
                continue
            revenue_value = to_float(revenue)
            if revenue_value is None:
                continue
            month_num = int(month_match.group(1))
            revenue_trend.append({"month": f"2026-{month_num:02d}", "revenue": revenue_value})

        for sheet_name, occupancy, revenue, net_profit_value, margin in zip(
            month_cells, occupancy_cells, revenue_cells, net_profit_cells, margin_cells
        ):
            if not sheet_name:
                continue
            month_match = re.search(r"①R\d+_(\d+)集計", str(sheet_name))
            if not month_match:
                continue
            month_key = f"2026-{int(month_match.group(1)):02d}"
            if month_key == PREVIOUS_MONTH:
                previous_revenue = to_float(revenue)
                previous_occupancy_pct = to_float(occupancy)
                previous_net_profit = to_float(net_profit_value)
                previous_margin_pct = to_float(margin)

    if special_sheet_name:
        ws = workbook[special_sheet_name]
        for row in ws.iter_rows(min_row=7, values_only=True):
            special_cleaning_revenue += to_float(row_item(row, 3)) or 0.0
            special_cleaning_cost += to_float(row_item(row, 9)) or 0.0

    return {
        "region": infer_region(hotel_name),
        "monthlyRevenue": monthly_revenue,
        "previousRevenue": previous_revenue,
        "targetRevenue": target_revenue,
        "marginPct": margin_pct * 100 if margin_pct is not None else None,
        "previousMarginPct": previous_margin_pct * 100 if previous_margin_pct is not None else None,
        "targetMarginPct": target_margin_pct * 100 if target_margin_pct is not None else None,
        "occupancyPct": occupancy_pct * 100 if occupancy_pct is not None else None,
        "previousOccupancyPct": previous_occupancy_pct * 100 if previous_occupancy_pct is not None else None,
        "variableCostRate": variable_cost_rate * 100 if variable_cost_rate is not None else None,
        "netProfit": net_profit,
        "previousNetProfit": previous_net_profit,
        "claimRate": claim_rate * 100 if claim_rate is not None else None,
        "roomsCleanedTodaySample": rooms_cleaned,
        "daysInMonth": days_in_month,
        "specialCleaningRevenue": special_cleaning_revenue or None,
        "specialCleaningCost": special_cleaning_cost or None,
        "revenueTrend": revenue_trend,
    }


def build_hotel_record(hotel_name: str, reviews: list[dict[str, Any]], financials: dict[str, Any]) -> dict[str, Any]:
    ratings = [review["normalizedRating"] for review in reviews if review["normalizedRating"] is not None]
    avg_review = sum(ratings) / len(ratings) if ratings else None

    positive_categories = Counter(review["positiveCategory"] for review in reviews)
    negative_categories = Counter(review["negativeCategory"] for review in reviews)
    cleaning_negative_count = sum(
        1 for review in reviews if review["cleaningRelated"] and review["negativeCategory"] != "未分類"
    )
    cleaning_positive_count = sum(
        1 for review in reviews if review["cleaningRelated"] and review["positiveCategory"] != "未分類"
    )
    review_count = len(reviews)
    cleaning_negative_rate = (cleaning_negative_count / review_count * 100) if review_count else 0.0
    risk_level = infer_risk_level(avg_review, cleaning_negative_rate)
    upsell_level = infer_upsell_level(review_count, cleaning_negative_rate, financials["monthlyRevenue"])

    main_issue = top_category(
        Counter(
            review["negativeCategory"]
            for review in reviews
            if review["negativeCategory"] != "未分類" and review["cleaningRelated"]
        )
    )
    if main_issue == "特記事項なし":
        main_issue = top_category(negative_categories)

    main_praise = top_category(
        Counter(
            review["positiveCategory"]
            for review in reviews
            if review["positiveCategory"] != "未分類" and review["cleaningRelated"]
        )
    )
    if main_praise == "特記事項なし":
        main_praise = top_category(positive_categories)

    next_action = build_next_action(main_issue, risk_level, upsell_level)
    current_month_reviews = [review for review in reviews if str(review["reviewDate"]).startswith(CURRENT_MONTH)]
    previous_month_reviews = [review for review in reviews if str(review["reviewDate"]).startswith(PREVIOUS_MONTH)]

    def month_review_metrics(items: list[dict[str, Any]]) -> dict[str, Any]:
        month_ratings = [item["normalizedRating"] for item in items if item["normalizedRating"] is not None]
        month_avg_review = sum(month_ratings) / len(month_ratings) if month_ratings else None
        month_cleaning_negative_count = sum(
            1 for item in items if item["cleaningRelated"] and item["negativeCategory"] != "未分類"
        )
        month_cleaning_negative_rate = (
            month_cleaning_negative_count / len(items) * 100 if items else None
        )
        return {
            "reviewCount": len(items),
            "avgReview": round(month_avg_review, 1) if month_avg_review is not None else None,
            "cleaningNegativeRate": round(month_cleaning_negative_rate, 1)
            if month_cleaning_negative_rate is not None
            else None,
        }

    current_month_metrics = month_review_metrics(current_month_reviews)
    previous_month_metrics = month_review_metrics(previous_month_reviews)
    elapsed_days = CURRENT_DATE.day
    days_in_month = financials.get("daysInMonth") or calendar.monthrange(CURRENT_DATE.year, CURRENT_DATE.month)[1]
    projected_revenue = (
        financials["monthlyRevenue"] / elapsed_days * days_in_month
        if financials["monthlyRevenue"] is not None and elapsed_days > 0
        else None
    )
    target_progress_pct = (
        financials["monthlyRevenue"] / financials["targetRevenue"] * 100
        if financials["monthlyRevenue"] is not None and financials["targetRevenue"] not in (None, 0)
        else None
    )
    projected_target_attainment_pct = (
        projected_revenue / financials["targetRevenue"] * 100
        if projected_revenue is not None and financials["targetRevenue"] not in (None, 0)
        else None
    )

    return {
        "name": hotel_name,
        "region": financials["region"],
        "monthlyRevenue": financials["monthlyRevenue"],
        "previousRevenue": financials["previousRevenue"],
        "targetRevenue": financials["targetRevenue"],
        "marginPct": round(financials["marginPct"], 1) if financials["marginPct"] is not None else None,
        "previousMarginPct": round(financials["previousMarginPct"], 1)
        if financials["previousMarginPct"] is not None
        else None,
        "targetMarginPct": round(financials["targetMarginPct"], 1)
        if financials["targetMarginPct"] is not None
        else None,
        "occupancyPct": round(financials["occupancyPct"], 1) if financials["occupancyPct"] is not None else None,
        "previousOccupancyPct": round(financials["previousOccupancyPct"], 1)
        if financials["previousOccupancyPct"] is not None
        else None,
        "variableCostRate": round(financials["variableCostRate"], 1)
        if financials["variableCostRate"] is not None
        else None,
        "netProfit": financials["netProfit"],
        "previousNetProfit": financials["previousNetProfit"],
        "claimRate": round(financials["claimRate"], 2) if financials["claimRate"] is not None else None,
        "specialCleaningRevenue": financials["specialCleaningRevenue"],
        "specialCleaningCost": financials["specialCleaningCost"],
        "avgReview": round(avg_review, 1) if avg_review is not None else None,
        "reviewCount": review_count,
        "cleaningNegativeRate": round(cleaning_negative_rate, 1),
        "cleaningPositiveCount": cleaning_positive_count,
        "cleaningNegativeCount": cleaning_negative_count,
        "mainIssue": main_issue,
        "mainPraise": main_praise,
        "riskLevel": risk_level,
        "upsellLevel": upsell_level,
        "nextAction": next_action,
        "revenueTrend": financials["revenueTrend"],
        "elapsedDays": elapsed_days,
        "daysInMonth": days_in_month,
        "projectedRevenue": projected_revenue,
        "targetProgressPct": round(target_progress_pct, 1) if target_progress_pct is not None else None,
        "projectedTargetAttainmentPct": round(projected_target_attainment_pct, 1)
        if projected_target_attainment_pct is not None
        else None,
        "currentMonthMetrics": current_month_metrics,
        "previousMonthMetrics": previous_month_metrics,
        "revenueDelta": (
            financials["monthlyRevenue"] - financials["previousRevenue"]
            if financials["monthlyRevenue"] is not None and financials["previousRevenue"] is not None
            else None
        ),
        "marginDelta": (
            round(financials["marginPct"] - financials["previousMarginPct"], 1)
            if financials["marginPct"] is not None and financials["previousMarginPct"] is not None
            else None
        ),
        "occupancyDelta": (
            round(financials["occupancyPct"] - financials["previousOccupancyPct"], 1)
            if financials["occupancyPct"] is not None and financials["previousOccupancyPct"] is not None
            else None
        ),
        "reviewDelta": (
            round(current_month_metrics["avgReview"] - previous_month_metrics["avgReview"], 1)
            if current_month_metrics["avgReview"] is not None and previous_month_metrics["avgReview"] is not None
            else None
        ),
        "cleaningNegativeRateDelta": (
            round(
                current_month_metrics["cleaningNegativeRate"] - previous_month_metrics["cleaningNegativeRate"], 1
            )
            if current_month_metrics["cleaningNegativeRate"] is not None
            and previous_month_metrics["cleaningNegativeRate"] is not None
            else None
        ),
    }


def build_actions(hotels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    risk_weight = {"high": 2, "medium": 1, "low": 0}
    sorted_hotels = sorted(
        [
            hotel
            for hotel in hotels
            if hotel["riskLevel"] != "low" or hotel["upsellLevel"] == "high"
        ],
        key=lambda hotel: (
            -risk_weight[hotel["riskLevel"]],
            -(hotel["monthlyRevenue"] or 0),
            -hotel["cleaningNegativeRate"],
        ),
    )

    actions = []
    for hotel in sorted_hotels[:8]:
        if hotel["riskLevel"] == "high":
            reason = (
                f"評価 {hotel['avgReview'] or '未取得'} / 清掃ネガ率 {hotel['cleaningNegativeRate']}%"
                + (
                    f" / 前月差 {hotel['cleaningNegativeRateDelta']}pt"
                    if hotel["cleaningNegativeRateDelta"] is not None
                    else ""
                )
            )
            owner = "Operations"
            priority = "High"
        elif hotel["upsellLevel"] == "high":
            reason = f"月商 {int(hotel['monthlyRevenue'] or 0):,}円 / 次提案: {hotel['mainIssue']}"
            owner = "Sales"
            priority = "Medium"
        else:
            reason = f"品質監視継続 / 主課題: {hotel['mainIssue']}"
            owner = "Management"
            priority = "Low"

        actions.append(
            {
                "priority": priority,
                "hotel": hotel["name"],
                "reason": reason,
                "owner": owner,
            }
        )
    return actions


def combine_revenue_trend(hotels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[str, float] = {}
    for hotel in hotels:
        for item in hotel.get("revenueTrend", []):
            totals[item["month"]] = totals.get(item["month"], 0.0) + float(item["revenue"])
    return [{"month": month, "revenue": round(totals[month], 1)} for month in sorted(totals.keys())]


def build_executive_summary(summary: dict[str, Any], hotels: list[dict[str, Any]]) -> list[str]:
    under_target = sorted(
        [
            hotel
            for hotel in hotels
            if hotel["projectedTargetAttainmentPct"] is not None and hotel["projectedTargetAttainmentPct"] < 90
        ],
        key=lambda hotel: hotel["projectedTargetAttainmentPct"],
    )
    high_risk = sorted(
        [hotel for hotel in hotels if hotel["riskLevel"] == "high"],
        key=lambda hotel: (
            -(hotel["cleaningNegativeRateDelta"] or 0),
            -(hotel["cleaningNegativeRate"] or 0),
            -(hotel["monthlyRevenue"] or 0),
        ),
    )
    strong_candidates = sorted(
        [hotel for hotel in hotels if hotel["upsellLevel"] == "high"],
        key=lambda hotel: ((hotel["projectedTargetAttainmentPct"] or 0), -(hotel["monthlyRevenue"] or 0)),
        reverse=True,
    )

    line1 = (
        f"3月13日時点の売上は {round((summary['monthlyRevenue'] or 0) / 10000):,}万円、"
        f"月末着地見込みは {round((summary['projectedMonthlyRevenue'] or 0) / 10000):,}万円です。"
        f"平均達成見込みは {summary['avgProjectedTargetAttainmentPct'] or '未連携'}% です。"
    )

    if high_risk:
        targets = "、".join(hotel["name"] for hotel in high_risk[:3])
        line2 = (
            f"品質面では {targets} の優先度が高く、"
            f"清掃ネガ率や評価差の悪化を先に確認したい状況です。"
        )
    else:
        line2 = "品質面では大きな悪化ホテルは少なく、維持運用を中心に見ていけます。"

    if under_target:
        targets = "、".join(hotel["name"] for hotel in under_target[:3])
        line3 = (
            f"売上面では {targets} が90%未達見込みで、"
            f"進捗フォローか追加提案の優先対象です。"
        )
    elif strong_candidates:
        targets = "、".join(hotel["name"] for hotel in strong_candidates[:3])
        line3 = f"営業面では {targets} がアップセル候補で、特別清掃や品質提案の余地があります。"
    else:
        line3 = "営業面では大きな未達は少なく、既存顧客の深耕提案を進めやすい状況です。"

    return [line1, line2, line3]


def build_summary(hotels: list[dict[str, Any]], reviews: list[dict[str, Any]], revenue_trend: list[dict[str, Any]]) -> dict[str, Any]:
    revenue_values = [hotel["monthlyRevenue"] for hotel in hotels if hotel["monthlyRevenue"] is not None]
    special_values = [
        hotel["specialCleaningRevenue"] for hotel in hotels if hotel["specialCleaningRevenue"] is not None
    ]
    avg_review_values = [hotel["avgReview"] for hotel in hotels if hotel["avgReview"] is not None]

    current_revenue = sum(revenue_values) if revenue_values else None
    revenue_delta_pct = None
    if len(revenue_trend) >= 2 and revenue_trend[-2]["revenue"]:
        current = revenue_trend[-1]["revenue"]
        previous = revenue_trend[-2]["revenue"]
        revenue_delta_pct = (current - previous) / previous * 100

    cleaning_negative_total = sum(hotel["cleaningNegativeCount"] for hotel in hotels)
    total_reviews = len(reviews)
    previous_revenue_values = [hotel["previousRevenue"] for hotel in hotels if hotel["previousRevenue"] is not None]
    review_deltas = [hotel["reviewDelta"] for hotel in hotels if hotel["reviewDelta"] is not None]
    cleaning_rate_deltas = [
        hotel["cleaningNegativeRateDelta"]
        for hotel in hotels
        if hotel["cleaningNegativeRateDelta"] is not None
    ]
    improved_hotels = [
        hotel
        for hotel in hotels
        if hotel["reviewDelta"] is not None and hotel["reviewDelta"] > 0
    ]
    worsened_hotels = [
        hotel
        for hotel in hotels
        if (hotel["reviewDelta"] is not None and hotel["reviewDelta"] < 0)
        or (hotel["cleaningNegativeRateDelta"] is not None and hotel["cleaningNegativeRateDelta"] > 0)
    ]

    projected_revenues = [hotel["projectedRevenue"] for hotel in hotels if hotel["projectedRevenue"] is not None]
    target_progress_values = [
        hotel["targetProgressPct"] for hotel in hotels if hotel["targetProgressPct"] is not None
    ]
    projected_target_values = [
        hotel["projectedTargetAttainmentPct"]
        for hotel in hotels
        if hotel["projectedTargetAttainmentPct"] is not None
    ]

    return {
        "monthlyRevenue": round(current_revenue, 1) if current_revenue is not None else None,
        "previousMonthlyRevenue": round(sum(previous_revenue_values), 1) if previous_revenue_values else None,
        "projectedMonthlyRevenue": round(sum(projected_revenues), 1) if projected_revenues else None,
        "revenueDeltaPct": round(revenue_delta_pct, 1) if revenue_delta_pct is not None else None,
        "activeHotels": len(hotels),
        "atRiskHotels": sum(1 for hotel in hotels if hotel["riskLevel"] == "high"),
        "upsellCandidates": sum(1 for hotel in hotels if hotel["upsellLevel"] == "high"),
        "specialCleaningRevenue": round(sum(special_values), 1) if special_values else None,
        "avgReviewScore": round(sum(avg_review_values) / len(avg_review_values), 1) if avg_review_values else None,
        "cleaningNegativeRate": round(cleaning_negative_total / total_reviews * 100, 1) if total_reviews else 0.0,
        "avgReviewDelta": round(sum(review_deltas) / len(review_deltas), 1) if review_deltas else None,
        "cleaningNegativeRateDelta": round(sum(cleaning_rate_deltas) / len(cleaning_rate_deltas), 1)
        if cleaning_rate_deltas
        else None,
        "improvedHotels": len(improved_hotels),
        "worsenedHotels": len(worsened_hotels),
        "elapsedDays": CURRENT_DATE.day,
        "daysInMonth": calendar.monthrange(CURRENT_DATE.year, CURRENT_DATE.month)[1],
        "avgTargetProgressPct": round(sum(target_progress_values) / len(target_progress_values), 1)
        if target_progress_values
        else None,
        "avgProjectedTargetAttainmentPct": round(sum(projected_target_values) / len(projected_target_values), 1)
        if projected_target_values
        else None,
    }


def main() -> None:
    master_rows = fetch_csv_rows(MASTER_EXPORT_URL)
    report_rows = [row for row in master_rows[1:] if len(row) >= 2 and row[0] and row[1]]

    hotels: list[dict[str, Any]] = []
    reviews: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []

    for row in report_rows:
        hotel_name = (row[2] if len(row) > 2 and row[2] else row[0]).strip()
        report_url = row[1].strip()
        sheet_id = extract_sheet_id(report_url)

        if not sheet_id:
            skipped.append({"hotel": hotel_name, "reason": "sheet_id_not_found"})
            continue

        print(f"Fetching {hotel_name}...")
        try:
            review_csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            review_rows = fetch_csv_rows(review_csv_url)
            hotel_reviews = parse_reviews(review_rows, hotel_name)

            xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            workbook = load_workbook(io.BytesIO(fetch_bytes(xlsx_url)), data_only=True, read_only=True)
            financials = extract_financials(workbook, hotel_name)

            reviews.extend(hotel_reviews)
            hotels.append(build_hotel_record(hotel_name, hotel_reviews, financials))
        except Exception as error:  # noqa: BLE001
            skipped.append({"hotel": hotel_name, "reason": str(error)})

    revenue_trend = combine_revenue_trend(hotels)
    hotels = sorted(hotels, key=lambda hotel: (hotel["monthlyRevenue"] or 0), reverse=True)
    payload = {
        "updatedAt": datetime.now().astimezone().strftime("%Y/%m/%d %H:%M:%S"),
        "summary": build_summary(hotels, reviews, revenue_trend),
        "revenueTrend": revenue_trend,
        "hotels": hotels,
        "actions": build_actions(hotels),
        "executiveSummary": [],
        "metadata": {
            "sourceSheetId": MASTER_SHEET_ID,
            "sourceGid": MASTER_GID,
            "reviewCount": len(reviews),
            "skippedHotels": skipped,
            "note": "口コミと各ホテルブック内の月報・集計・年間集計・特別清掃シートを使用",
        },
    }
    payload["executiveSummary"] = build_executive_summary(payload["summary"], hotels)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "generated-dashboard.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "live-data.js").write_text(
        f"window.primeChangeDataLive = {json.dumps(payload, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )
    print(f"Generated dashboard data for {len(hotels)} hotels.")


if __name__ == "__main__":
    main()
